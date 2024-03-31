import os
import openpyxl
import csv
import datetime

# 定义字母和数字的映射表
letter_to_number = {
    'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4,
    'F': 5, 'G': 6, 'H': 7, 'I': 8, 'J': 9,
    'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14,
    'P': 15, 'Q': 16, 'R': 17, 'S': 18, 'T': 19,
    'U': 20, 'V': 21, 'W': 22, 'X': 23, 'Y': 24,
    'Z': 25, 'AA': 26,'AB': 27,'AC': 28,'AD': 29,
    'AE': 30,'AF': 31,'AG': 32,'AH': 33,'AI': 34,
    'AJ': 35,'AK': 36,
}

# 反转字母和数字的映射表，以便进行数字到字母的转换
number_to_letter = {v: k for k, v in letter_to_number.items()}
# 测试映射表
# print(letter_to_number['A'])  # 输出: 0
# print(number_to_letter[0])    # 输出: 'A'


def copyCSVtoXlsx(csv_directory,csv_filename,excel_file):
    csv_file = os.path.join(csv_directory, csv_filename)
    excel_file = os.path.join(csv_directory, excel_file)
    sheet_name = os.path.splitext(csv_filename)[0]

    if not os.path.exists(csv_directory):
        os.makedirs(csv_directory)

    # 创建或加载Excel工作簿
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
    else:
        workbook = openpyxl.Workbook()
        
    # 读取CSV文件
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader) 
        data = list(reader)
        
    # 创建或加载工作表
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(title=sheet_name)


    # 将数据写入工作表
    for row_idx, row_data in enumerate(data, start=2):
        # 检查第一列是否为“合计”，如果是，则跳过当前行
        if row_data[0] == "合计":
            continue
        for col_idx, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 保存工作簿
    workbook.save(excel_file)
    print("CSV文件已成功复制到Excel文件的对应工作表中。")

def parse_datetime(input_datetime, row_idx=None):
    if isinstance(input_datetime, datetime.datetime):
        return input_datetime
    elif isinstance(input_datetime, str):
        try:
            return datetime.datetime.strptime(input_datetime, "%Y-%m-%d")
        except ValueError as e:
            if row_idx is not None:
                print(f"日期解析错误：{e}，位于第 {row_idx} 行")
            else:
                print(f"日期解析错误：{e}")
                if row_idx is not None:
                    print(f"类型错误：输入必须是字符串或 datetime.datetime 对象，位于第 {row_idx} 行")
                else:
                    print("类型错误：输入必须是字符串或 datetime.datetime 对象")
            raise TypeError("Input must be a string or datetime.datetime object")

#整体 
def processXlsx_ShuiHouHuiKuan(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")
    
    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取收款明细表中M列的求和
        sum_column = 0
        
        if "收款明细表" in workbook.sheetnames:
            worksheet_receipt = workbook["收款明细表"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_date_str = row[letter_to_number['K']]  # K列，收款时间字符串
                receipt_amount = row[letter_to_number['M']]  # M列，收款金额
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio/10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_ZhiXiaoYingShouQueBao(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中Q列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['Q']]  # Q列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '直销' and receipt_is_amount == '未回款':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_FenXiaoYingShouQueBao(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中Q列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['Q']]  # Q列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '分销' and receipt_is_amount == '未回款':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_XinQianQueBao(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取项目漏斗汇总-签约金额替重中Y列的求和
        sum_column = 0
        if "项目漏斗汇总-签约金额替重" in workbook.sheetnames:
            worksheet_receipt = workbook["项目漏斗汇总-签约金额替重"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_date_str = row[letter_to_number['X']]  # X列，日期
                receipt_amount = row[letter_to_number['Y']]  # Y列，回款

                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_ZhiXiaoYingShouChongCi(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中R列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['R']]  # R列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '直销' and receipt_is_amount == '未回款':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_FenXiaoYingShouChongCi(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中R列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['R']]  # R列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '分销' and receipt_is_amount == '未回款':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_XinQianChongCi(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取项目漏斗汇总-签约金额替重中Z列的求和
        sum_column = 0
        if "项目漏斗汇总-签约金额替重" in workbook.sheetnames:
            worksheet_receipt = workbook["项目漏斗汇总-签约金额替重"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_date_str = row[letter_to_number['X']]  # X列，日期
                receipt_amount = row[letter_to_number['Z']]  # Z列，冲刺回款

                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

#直销
def processXlsx_ShuiHouHuiKuan_ZhiXiao(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")
    
    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取收款明细表中M列的求和
        sum_column = 0
        if "收款明细表" in workbook.sheetnames:
            worksheet_receipt = workbook["收款明细表"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['I']]  # I列，直销分销
                receipt_date_str = row[letter_to_number['K']]  # K列，收款时间字符串
                receipt_amount = row[letter_to_number['M']]  # M列，收款金额
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales=='直销':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio/10000.0
        worksheet_24[unit] = sum_column
    return workbook
#分销
def processXlsx_ShuiHouHuiKuan_FenXiao(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")
    
    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取收款明细表中M列的求和
        sum_column = 0
        if "收款明细表" in workbook.sheetnames:
            worksheet_receipt = workbook["收款明细表"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['I']]  # I列，直销分销
                receipt_date_str = row[letter_to_number['K']]  # K列，收款时间字符串
                receipt_amount = row[letter_to_number['M']]  # M列，收款金额
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales=='分销':
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio/10000.0
        worksheet_24[unit] = sum_column
    return workbook

#订阅
def processXlsx_ShuiHouHuiKuan_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")
    
    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        sum_column = 0
        if "收款明细表" in workbook.sheetnames:
            worksheet_receipt = workbook["收款明细表"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_booking = row[letter_to_number['J']]  # J列，订阅与否
                receipt_date_str = row[letter_to_number['K']]  # K列，收款时间字符串
                receipt_amount = row[letter_to_number['M']]  # M列，收款金额
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and '非订阅' not in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '收款明细表' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio/10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_ZhiXiaoYingShouQueBao_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中Q列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_booking = row[letter_to_number['K']]  # K列，订阅与否
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['Q']]  # Q列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '直销' and receipt_is_amount == '未回款'and '非订阅' not in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_FenXiaoYingShouQueBao_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中Q列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_booking = row[letter_to_number['K']]  # K列，订阅与否
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['Q']]  # Q列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '分销' and receipt_is_amount == '未回款'and '非订阅' not in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_XinQianQueBao_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取项目漏斗汇总-签约金额替重中Y列的求和
        sum_column = 0
        if "项目漏斗汇总-签约金额替重" in workbook.sheetnames:
            worksheet_receipt = workbook["项目漏斗汇总-签约金额替重"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_booking = row[letter_to_number['W']]  # W列，订阅与否
                receipt_date_str = row[letter_to_number['X']]  # X列，日期
                receipt_amount = row[letter_to_number['Y']]  # Y列，回款

                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and '订阅' in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_ZhiXiaoYingShouChongCi_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中R列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_booking = row[letter_to_number['K']]  # K列，订阅与否
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['R']]  # R列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '直销' and receipt_is_amount == '未回款'and '非订阅' not in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_FenXiaoYingShouChongCi_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取应收及分销预测汇总中R列的求和
        sum_column = 0
        if "应收及分销预测汇总" in workbook.sheetnames:
            worksheet_receipt = workbook["应收及分销预测汇总"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_sales = row[letter_to_number['J']]  # J列，直销分销
                receipt_booking = row[letter_to_number['K']]  # K列，订阅与否
                receipt_date_str = row[letter_to_number['P']]  # P列，日期
                receipt_amount = row[letter_to_number['R']]  # R列，收款金额
                receipt_is_amount = row[letter_to_number['V']]  # V列，是否未回款
                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and receipt_sales == '分销' and receipt_is_amount == '未回款'and '非订阅' not in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook

def processXlsx_XinQianChongCi_DingYue(workbook, unit, dateStart, dateEnd,ratio,product_name):
    # 解析整数形式的日期为日期对象
    receipt_date_start = datetime.datetime.strptime(str(dateStart), "%Y%m%d")
    receipt_date_end = datetime.datetime.strptime(str(dateEnd), "%Y%m%d")

    if "24年业绩预测-机构" in workbook.sheetnames:
        worksheet_24 = workbook["24年业绩预测-机构"]
        # 获取项目漏斗汇总-签约金额替重中Z列的求和
        sum_column = 0
        if "项目漏斗汇总-签约金额替重" in workbook.sheetnames:
            worksheet_receipt = workbook["项目漏斗汇总-签约金额替重"]
            for row_idx, row in enumerate(worksheet_receipt.iter_rows(min_row=2, min_col=letter_to_number['A'], max_col=letter_to_number['AF'], values_only=True), start=2):
                # 检查是否为空行
                if not any(row):
                    continue
                receipt_booking = row[letter_to_number['W']]  # W列，订阅与否
                receipt_date_str = row[letter_to_number['X']]  # X列，日期
                receipt_amount = row[letter_to_number['Z']]  # Z列，冲刺回款

                try:
                    receipt_date = parse_datetime(receipt_date_str, row_idx=row_idx)
                    if receipt_date_start <= receipt_date <= receipt_date_end and '订阅' in receipt_booking:
                        if isinstance(receipt_amount, str):  # 如果值是字符串类型
                            # 删除逗号并尝试将字符串转换为浮点数
                            try:
                                cleaned_value = receipt_amount.replace(',', '')
                                # 如果字符串包含负号，我们在转换之前将其删除
                                if '-' in cleaned_value:
                                    cleaned_value = cleaned_value.replace('-', '')
                                    sum_column -= float(cleaned_value)
                                else:
                                    sum_column += float(cleaned_value)
                            except ValueError as e:
                                # 如果无法转换为浮点数，则打印错误消息并继续下一个值
                                print(f"数值错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
                        elif isinstance(receipt_amount, (int, float)):  # 如果值是数值类型
                            sum_column += receipt_amount
                except ValueError as e:
                    print(f"日期解析错误：{e}，位于工作表 '应收及分销预测汇总' 的第 {row_idx} 行")
        # 在D4单元格填充求和值
        sum_column = sum_column / ratio / 10000.0
        worksheet_24[unit] = sum_column
    return workbook


